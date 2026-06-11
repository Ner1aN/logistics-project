from io import BytesIO

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill('solid', fgColor='274C77')
SUMMARY_FILL = PatternFill('solid', fgColor='EAF3F1')
WHITE_FONT = Font(color='FFFFFF', bold=True)
BOLD_FONT = Font(bold=True)


def _money(value):
    return float(value or 0)


def _number(value):
    return float(value or 0)


def _date(value):
    return value.strftime('%d.%m.%Y') if value else ''


def _request_transportation_value(request, attr, default=''):
    transportation = getattr(request, 'transportation', None)
    if not transportation:
        return default
    value = getattr(transportation, attr)
    return value() if callable(value) else value


def _autosize_columns(sheet):
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = '' if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 45)


def _style_header(row):
    for cell in row:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')


def _append_table(sheet, headers, rows):
    sheet.append(headers)
    _style_header(sheet[sheet.max_row])
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = 'A2'
    sheet.auto_filter.ref = sheet.dimensions
    _autosize_columns(sheet)


def build_reports_excel(context, filters):
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = 'Сводка'

    summary_sheet.append(['Отчет по грузоперевозкам'])
    summary_sheet['A1'].font = Font(size=16, bold=True)
    summary_sheet.append(['Дата формирования', timezone.localtime().strftime('%d.%m.%Y %H:%M')])
    summary_sheet.append([])
    summary_sheet.append(['Фильтр', 'Значение'])
    _style_header(summary_sheet[4])

    filter_rows = [
        ('С даты', filters.get('date_from') or 'Все'),
        ('По дату', filters.get('date_to') or 'Все'),
        ('Статус', filters.get('status') or 'Все'),
        ('Клиент ID', filters.get('client') or 'Все'),
        ('Тип груза', filters.get('cargo_type') or 'Все'),
    ]
    for row in filter_rows:
        summary_sheet.append(row)

    summary_sheet.append([])
    summary_sheet.append(['Показатель', 'Значение'])
    _style_header(summary_sheet[summary_sheet.max_row])
    metrics = [
        ('Заявок в отчете', context['request_total_count']),
        ('Сумма по заявкам', _money(context['request_total_cost'])),
        ('Средняя стоимость', _money(context['average_request_cost'])),
        ('Выполнено заявок, %', context['completed_request_percent']),
        ('Просроченные заявки', context['overdue_request_count']),
        ('Активные в срок', context['active_request_count']),
        ('Выполненные перевозки', context['completed_transportation_count']),
        ('Фактические рейсы ТС', context['completed_transportation_trips']),
        ('Километраж выполненных перевозок, км', _number(context['completed_transportation_distance'])),
        ('Выручка по выполненным', _money(context['completed_transportation_cost'])),
    ]
    for row in metrics:
        summary_sheet.append(row)

    for row in summary_sheet.iter_rows(min_row=5):
        for cell in row:
            cell.alignment = Alignment(vertical='center')
    for row_number in range(10, 20):
        summary_sheet[f'A{row_number}'].fill = SUMMARY_FILL
        summary_sheet[f'B{row_number}'].fill = SUMMARY_FILL
    _autosize_columns(summary_sheet)

    status_sheet = workbook.create_sheet('Статусы')
    _append_table(
        status_sheet,
        ['Статус', 'Количество', 'Сумма'],
        [
            [row['status__name'], row['total'], _money(row['total_cost'])]
            for row in context['status_stats']
        ],
    )

    requests_sheet = workbook.create_sheet('Заявки')
    _append_table(
        requests_sheet,
        ['ID', 'Клиент', 'Тип груза', 'Описание груза', 'Маршрут', 'Рейсов ТС', 'Километраж, км', 'Дата', 'Статус', 'Стоимость'],
        [
            [
                request.pk,
                request.client.name,
                request.get_cargo_type_display(),
                request.cargo_name,
                f'{request.route_from} -> {request.route_to}',
                _request_transportation_value(request, 'trip_count'),
                _number(_request_transportation_value(request, 'total_distance_km', 0)),
                _date(request.transportation_date),
                request.status.name,
                _money(request.cost),
            ]
            for request in context['requests']
        ],
    )

    transportations_sheet = workbook.create_sheet('Выполненные перевозки')
    _append_table(
        transportations_sheet,
        [
            'Заявка',
            'Клиент',
            'Водитель',
            'Транспорт',
            'Рейсов ТС',
            'Стоянка-погрузка, км',
            'Погрузка-заказчик, км',
            'Заказчик-погрузка, км',
            'Общий километраж, км',
            'Стоимость километража',
            'Дата перевозки',
            'Дата завершения',
            'Стоимость',
        ],
        [
            [
                item.request.pk,
                item.request.client.name,
                item.driver.full_name,
                item.vehicle.registration_number,
                item.trip_count,
                _number(item.distance_parking_to_loading_km),
                _number(item.distance_loading_to_customer_km),
                _number(item.distance_customer_to_loading_km),
                _number(item.total_distance_km),
                _money(item.distance_cost),
                _date(item.request.transportation_date),
                _date(item.arrival_at or item.request.transportation_date),
                _money(item.request.cost),
            ]
            for item in context['completed_transportations']
        ],
    )

    drivers_sheet = workbook.create_sheet('Загрузка водителей')
    _append_table(
        drivers_sheet,
        ['Водитель', 'Выполненных перевозок', 'Фактических рейсов'],
        [
            [row['driver__full_name'], row['total'], row['trip_total'] or 0]
            for row in context['driver_load_stats']
        ],
    )

    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()
